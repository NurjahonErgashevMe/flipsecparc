#!/usr/bin/env python3
"""
Экспорт house_pages_result.json в Excel (.xlsx).

Вход: JSON-массив объектов из house_pages_parser.py.
Выход: xlsx с 15 колонками (14 полей + house_id).
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Нужен пакет openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

_DIR = Path(__file__).resolve().parent

# (Заголовок в Excel, ключ в JSON)
SCHEMA: list[tuple[str, str]] = [
    ("Адрес (жилой дом, Москва)", "address"),
    ("Год постройки", "year"),
    ("Перекрытия", "overlaps"),
    ("Тип дома", "house_type"),
    ("Этаж / этажность", "floors_text"),
    ("Серия дома", "series"),
    ("Округ", "okrug"),
    ("Район", "rayon"),
    ("Лифты (пасс., груз.)", "lifts"),
    ("Высота потолков", "ceiling_height"),
    ("Управляющая компания", "management_company"),
    ("Жилой комплекс", "residential_complex"),
    ("Метро (пешком / транспорт)", "metro"),
    ("Застройщик", "developer"),
    ("ID дома", "house_id"),
]

WIDTHS: dict[str, float] = {
    "address": 54,
    "year": 12,
    "overlaps": 18,
    "house_type": 24,
    "floors_text": 20,
    "series": 32,
    "okrug": 18,
    "rayon": 22,
    "lifts": 24,
    "ceiling_height": 16,
    "management_company": 38,
    "residential_complex": 28,
    "metro": 40,
    "developer": 24,
    "house_id": 12,
}


def load_items(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8")
    data = json.loads(raw)
    if not isinstance(data, list):
        raise ValueError("Ожидается JSON-массив объектов домов")
    return [item for item in data if isinstance(item, dict)]


def cell_value(v: Any) -> Any:
    if v is None:
        return ""
    return v


def main() -> None:
    parser = argparse.ArgumentParser(description="house_pages_result.json -> Excel (.xlsx)")
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=_DIR / "house_pages_result.json",
        help="Входной JSON (по умолчанию house_pages_result.json рядом со скриптом)",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Выходной .xlsx (по умолчанию: имя входного файла с расширением .xlsx)",
    )
    args = parser.parse_args()

    in_path: Path = args.input
    if not in_path.is_file():
        print(f"Файл не найден: {in_path}", file=sys.stderr)
        raise SystemExit(2)

    out_path: Path = args.output or in_path.with_suffix(".xlsx")
    items = load_items(in_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "houses_pages"

    header_font = Font(bold=True)
    header_align = Alignment(wrap_text=True, vertical="center")
    top = Alignment(vertical="top")
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [title for title, _ in SCHEMA]
    keys = [key for _, key in SCHEMA]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align

    wrapped_keys = {"address", "series", "management_company", "metro"}
    wrapped_cols = {keys.index(k) + 1 for k in wrapped_keys}

    for row_idx, item in enumerate(items, start=2):
        for col_idx, key in enumerate(keys, start=1):
            val = cell_value(item.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap if col_idx in wrapped_cols else top

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{len(items) + 1}"

    for i, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(key, 14)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Записано строк: {len(items)}, файл: {out_path.resolve()}")


if __name__ == "__main__":
    main()
