#!/usr/bin/env python3
"""
Экспорт result.json в Excel (.xlsx).

Режим по умолчанию: только жилые дома Москвы и 15 колонок по ТЗ.
Часть полей в исходном API flatinfo отсутствует — колонки остаются, ячейки пустые.

Требуется: pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Нужен пакет openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

_DIR = Path(__file__).resolve().parent

ALLOWED_CITY_IDS = {"1", "12552", "12565"}
DISALLOWED_STREET_IDS = {"3745"}

# (заголовок в Excel, внутренний ключ строки)
BRIEF_SCHEMA: list[tuple[str, str]] = [
    ("Адрес (жилой дом, Москва)", "address"),
    ("Год постройки", "year"),
    ("Перекрытия", "overlaps"),  # нет в flatinfo JSON
    ("Тип дома", "house_type"),
    ("Этаж / этажность", "floors_text"),
    ("Серия дома", "series"),
    ("Округ", "okrug"),  # нет в JSON
    ("Район", "rayon"),  # нет в JSON
    ("Лифты (пасс., груз.)", "lifts"),
    ("Высота потолков", "ceiling"),  # нет в JSON
    ("Управляющая компания", "uk"),  # нет в JSON
    ("Жилой комплекс", "jk"),
    ("Метро (пешком / транспорт)", "metro"),  # нет в JSON
    ("Застройщик", "developer"),  # нет в JSON
    ("ID дома", "house_id"),
]

LEGACY_COLUMNS: list[str] = [
    "house_id",
    "city",
    "city_id",
    "street",
    "street_id",
    "jk_id",
    "jk_name",
    "ser_id",
    "ser_name",
    "subser_id",
    "subser_name",
    "house_num",
    "year",
    "flats",
    "podezd",
    "type",
    "type_id",
    "floor",
    "levels",
    "lift_p",
    "lift_g",
    "jil_type",
    "house_sales",
    "house_rents",
    "lat",
    "lng",
    "garbage",
]

LEGACY_WIDTHS: dict[str, float] = {
    "house_id": 10,
    "city": 18,
    "city_id": 8,
    "street": 36,
    "street_id": 10,
    "jk_id": 8,
    "jk_name": 24,
    "ser_id": 8,
    "ser_name": 28,
    "subser_id": 10,
    "subser_name": 24,
    "house_num": 12,
    "year": 8,
    "flats": 8,
    "podezd": 8,
    "type": 16,
    "type_id": 8,
    "floor": 14,
    "levels": 8,
    "lift_p": 8,
    "lift_g": 8,
    "jil_type": 12,
    "house_sales": 18,
    "house_rents": 18,
    "lat": 12,
    "lng": 12,
    "garbage": 10,
}

BRIEF_WIDTHS: dict[str, float] = {
    "address": 52,
    "year": 12,
    "overlaps": 14,
    "house_type": 16,
    "floors_text": 22,
    "series": 32,
    "okrug": 18,
    "rayon": 22,
    "lifts": 22,
    "ceiling": 14,
    "uk": 28,
    "jk": 28,
    "metro": 28,
    "developer": 22,
    "house_id": 12,
}


def _normalize_jil(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return value.replace("\u00a0", " ").replace("\u2009", " ").strip()


def matches_target(item: dict[str, Any]) -> bool:
    return str(item.get("city_id")) in ALLOWED_CITY_IDS and str(
        item.get("street_id")
    ) not in DISALLOWED_STREET_IDS


def is_moscow_residential(item: dict[str, Any]) -> bool:
    if _normalize_jil(item.get("jil_type")) != "Жилой":
        return False
    cid = str(item.get("city_id", "")).strip()
    city = (item.get("city") or "").strip()
    return cid == "1" or city == "Москва"


def _series_text(item: dict[str, Any]) -> str:
    ser = (item.get("ser_name") or "").strip()
    sub = (item.get("subser_name") or "").strip()
    if ser and sub and ser != sub:
        return f"{ser} ({sub})"
    return ser or sub


def brief_row(item: dict[str, Any]) -> dict[str, Any]:
    city = (item.get("city") or "").strip()
    street = (item.get("street") or "").strip()
    num = (item.get("house_num") or "").strip()
    parts = [p for p in (city, street, num) if p]
    address = ", ".join(parts)

    floor = (item.get("floor") or "").strip()
    levels = (item.get("levels") or "").strip()
    if floor and levels:
        floors_text = f"{floor} / этажность: {levels}"
    elif floor:
        floors_text = floor
    elif levels:
        floors_text = f"этажность: {levels}"
    else:
        floors_text = ""

    lp = item.get("lift_p", "")
    lg = item.get("lift_g", "")
    lifts = f"пасс.: {lp}; груз.: {lg}"

    hid = item.get("house_id")
    return {
        "address": address,
        "year": cell_value(item.get("year")),
        "overlaps": "",
        "house_type": cell_value(item.get("type")),
        "floors_text": floors_text,
        "series": _series_text(item),
        "okrug": "",
        "rayon": "",
        "lifts": lifts,
        "ceiling": "",
        "uk": "",
        "jk": cell_value(item.get("jk_name")),
        "metro": "",
        "developer": "",
        "house_id": hid if hid is not None else "",
    }


def load_items(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8")
    data = json.loads(raw)
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "items" in data:
        return list(data["items"])
    raise ValueError("Ожидается JSON-массив или объект с ключом «items»")


def cell_value(v: Any) -> Any:
    if v is None:
        return ""
    return v


def main() -> None:
    p = argparse.ArgumentParser(description="flatinfo result.json → Excel (.xlsx)")
    p.add_argument(
        "-i",
        "--input",
        type=Path,
        default=_DIR / "result.json",
        help="Входной JSON",
    )
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Выходной .xlsx",
    )
    p.add_argument(
        "--legacy-all-columns",
        action="store_true",
        help="Старый экспорт: все поля JSON (как раньше)",
    )
    p.add_argument(
        "--only-target",
        action="store_true",
        help="Доп. фильтр city_id / street_id (как в houses_parser)",
    )
    p.add_argument(
        "--no-moscow-filter",
        action="store_true",
        help="Не отфильтровывать только Москву+Жилой (для краткого режима)",
    )
    args = p.parse_args()
    in_path: Path = args.input
    if not in_path.is_file():
        print(f"Файл не найден: {in_path}", file=sys.stderr)
        raise SystemExit(2)
    out_path: Path = args.output or in_path.with_suffix(".xlsx")

    items = load_items(in_path)
    total = len(items)

    if args.legacy_all_columns:
        if args.only_target:
            items = [it for it in items if matches_target(it)]
            print(f"Всего в JSON: {total}, после --only-target: {len(items)}")
        columns = LEGACY_COLUMNS
        rows: list[Any] = items
        use_brief = False
    else:
        if not args.no_moscow_filter:
            items = [it for it in items if is_moscow_residential(it)]
        if args.only_target:
            items = [it for it in items if matches_target(it)]
        print(
            f"Всего в JSON: {total}, строк для Excel (Москва, жилой"
            + (", +only-target" if args.only_target else "")
            + f"): {len(items)}"
        )
        columns = BRIEF_SCHEMA
        rows = [brief_row(it) for it in items]
        use_brief = True

    wb = Workbook()
    ws = wb.active
    ws.title = "жилые_москва" if use_brief else "houses"

    header_font = Font(bold=True)
    top = Alignment(vertical="top")
    wrap = Alignment(wrap_text=True, vertical="top")
    header_align = Alignment(wrap_text=True, vertical="center")

    if use_brief:
        headers = [h for h, _ in BRIEF_SCHEMA]
        keys = [k for _, k in BRIEF_SCHEMA]
        for col_idx, title in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=title)
            c.font = header_font
            c.alignment = header_align
        addr_col = 1
        series_col = keys.index("series") + 1
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(keys, start=1):
                val = cell_value(row.get(key))
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx in (addr_col, series_col):
                    cell.alignment = wrap
                else:
                    cell.alignment = top
        ncols = len(headers)
        widths = BRIEF_WIDTHS
    else:
        for col_idx, name in enumerate(columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=name)
            c.font = header_font
            c.alignment = header_align
        ser_col = LEGACY_COLUMNS.index("ser_name") + 1
        subser_col = LEGACY_COLUMNS.index("subser_name") + 1
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(columns, start=1):
                val = cell_value(row.get(key))
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx in (ser_col, subser_col):
                    cell.alignment = wrap
                else:
                    cell.alignment = top
        ncols = len(columns)
        widths = LEGACY_WIDTHS

    last_row = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"

    if use_brief:
        for i, (_, key) in enumerate(BRIEF_SCHEMA, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)
    else:
        for i, name in enumerate(LEGACY_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Записано строк: {len(rows)}, файл: {out_path.resolve()}")


if __name__ == "__main__":
    main()
